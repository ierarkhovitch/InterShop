from django.core.management.base import BaseCommand, CommandError
from market.models import Category, Product
from prj.settings import DATA_DIR
from openpyxl import load_workbook


class Command(BaseCommand):

    def handle(self, *args, **options):
        print('Clear DB')
        Category.objects.all().delete()
        Product.objects.all().delete()

        print('Start import from excel %s' % DATA_DIR)
        wb = load_workbook((DATA_DIR + '\\price.xlsx'))
        worksheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        category = None
        for cnt in range(1, worksheet.max_row + 1):
            item = worksheet.cell(row=cnt, column=5).value
            cat = worksheet.cell(row=cnt, column=1).value
            if item == None:
                category = Category()
                category.name = cat
                category.save()
                print('Create category')
            else:
                if category:
                    product = Product()
                    product.name = item
                    product.category = category
                    product.save()
                    print('Create item')
